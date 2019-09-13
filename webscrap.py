import urllib.request
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup

def contaPaginas(pagina):
    soup = BeautifulSoup(page,"lxml")
    all_links = soup.find_all("dd")
    contagem = str(all_links[5])
    contagem = contagem.split("</")
    contagem = contagem[0]
    contagem = contagem.split(">")
    contagem = contagem[1]
    return contagem

# caminho = 'arquivo.xlsx'
# arquivo_excel = load_workbook(caminho)
# arquivo_excel.active
arquivo_excel = Workbook()


link = "http://192.168.254.41/general/information.html?kind=item"
page = urllib.request.urlopen(link,data=None)

planilha1 = arquivo_excel.active
planilha1.title = "impressoes"

max_linha = planilha1.max_row
print(max_linha)
colunaA = "A"+str(max_linha) 
colunaB = "B"+str(max_linha)
colunaC = "C"+str(max_linha)
planilha1[colunaA] = '192.168.254.40'
planilha1[colunaB] = '192.168.254.41'
planilha1[colunaC] = '192.168.254.42'



# valores = [
#     ("Categoria", "Valor"),
#     ("Restaurante", 45.99),
#     ("Transporte", 208.45),
#     ("Viagem", 558.54)
# ]
# for linha in valores:
#     planilha1.append(linha)

arquivo_excel.save("relatorio.xlsx")